"""XLSX ingestion (docs/28 §1.1).

Banks and payment processors ship `.xlsx` at least as often as `.csv`. Reads one
sheet into a string grid, then shares the header-detection / junk-row / row loop
with the CSV path (`ingest.tabular`).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from arbiter_engine.events.store import EventStore
from arbiter_engine.ingest.csv_source import IngestResult, _guard_duplicate, file_hash
from arbiter_engine.ingest.tabular import detect_header, ingest_rows, rows_from_grid
from arbiter_engine.specs.model import SourceSpec

MAX_BYTES = 50 * 1024 * 1024


def _cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _read_xlsx_grid(path: Path, sheet: str | None) -> list[list[str]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to read .xlsx files") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        return [[_cell(c) for c in row] for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def ingest_xlsx(
    store: EventStore,
    run_id: str,
    source_name: str,
    spec: SourceSpec,
    path: str | Path,
    *,
    profile: str | None = None,
    force: bool = False,
) -> IngestResult:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"source file not found: {p}")
    if p.stat().st_size > MAX_BYTES:
        raise ValueError(f"{p} exceeds the {MAX_BYTES // 1024 // 1024} MB cap")

    fh_hash = file_hash(p)
    _guard_duplicate(store, run_id, p.name, fh_hash, force)

    grid = _read_xlsx_grid(p, spec.sheet)
    header_idx = spec.header_row if spec.header_row is not None else detect_header(grid, spec)
    rows = rows_from_grid(grid, header_idx)

    rows_in, rows_ok, rows_q, pii, reasons = ingest_rows(
        store, run_id, source_name, spec, rows, fmt="xlsx", profile=profile, file_hash=fh_hash
    )
    return IngestResult(
        source=source_name,
        rows_in=rows_in,
        rows_ok=rows_ok,
        rows_quarantined=rows_q,
        pii_dropped=pii,
        file_hash=fh_hash,
        quarantine_reasons=reasons,
    )
